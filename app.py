import os, re, json, io, uuid
from flask import Flask, request, send_file, render_template, jsonify
from difflib import SequenceMatcher
import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(BASE_DIR, 'catalogue_cashmag.json'), 'r', encoding='utf-8') as f:
    CATALOGUE = json.load(f)

def normalize(s):
    s = (s or '').lower()
    s = s.replace('\u00e0','a').replace('\u00e2','a').replace('\u00e9','e').replace('\u00e8','e')
    s = s.replace('\u00ea','e').replace('\u00ee','i').replace('\u00f4','o').replace('\u00fb','u')
    s = s.replace('\u00f9','u').replace('\u00e7','c').replace("'",' ').replace('-',' ')
    return re.sub(r'\s+', ' ', s).strip()

def extract_volume(s):
    m = re.search(r'(\d+)\s*ml', (s or '').lower())
    return int(m.group(1)) if m else None

BRUIT = {'ultimate','sweet','edition','green','zero','classic','tabac','gourmand',
         'by','maison','le','pod','liquide','fizz','concentre','arome',
         'aromes','liquides','et','du','de','la','les','par','top','fill'}

def get_name_words(s):
    s = normalize(s)
    s = re.sub(r'^pack\s+|^concentre\s+', '', s)
    s = re.sub(r'\s*\d+\s*ml.*', '', s)
    for marque in ['aromes et liquides','a&l','fighter fuel','pulp','cupide','savourea',
                   'maison fuel','enfer','aspire','vaporesso','voopoo','geekvape',
                   'le french liquide','lemon time','fruizee','vampire vape','swoke',
                   'salt e vapor','juice heroes','liquideo','tjuice','lost vape','lostvape',
                   'polaris','moon shiners','le coq','coq qui vape','le petit verger',
                   'juice 66','hidden potion','panier du marche','my pulp','xfighter',
                   'smoke wars','e tasty','saiyen vapors','halo','french touch',
                   'tribal force','jnr','multi freeze','greenvillage','justfog','smoketech']:
        s = s.replace(normalize(marque), '')
    return [w for w in s.split() if len(w) > 2 and w not in BRUIT]

INDEX = {}
for i, p in enumerate(CATALOGUE):
    lib = normalize(p['libelle'])
    for w in set(lib.split()):
        if len(w) > 2:
            if w not in INDEX: INDEX[w] = []
            INDEX[w].append(i)

# Correspondances manuelles fournisseur → Cash Mag
MANUELS = {
    normalize('Lemon 50ml Lemon'): 'lemon time lemon',
    normalize("Lemon'time Lemon 50ml"): 'lemon time lemon',
    normalize('Tireboulette'): 'tireboulette',
    normalize('Tireboulette Peche Mangue Passion'): 'tireboulette',
}


# ── Table de correspondances résistances fournisseur → Cash Mag ───────────────
RESISTANCE_MAP = {
    # GeekVape B Series → RESISTANCE NANO
    'b series': 'resistance nano',
    'b boost': 'resistance nano',
    # GeekVape Z XM → RESISTANCE ZEUS SUB XM
    'z xm': 'resistance zeus sub',
    'zxm': 'resistance zeus sub',
    # GeekVape Zeus Sub → RESISTANCE ZEUS SUB
    'zeus subohm z series': 'resistance zeus sub',
    'zeus z series': 'resistance zeus sub',
    # Nautilus Aspire
    'nautilus': 'resistance nautilus',
    # PnP Voopoo
    'pnp x': 'resistance pnp',
    'pnp': 'resistance pnp',
    # Luxe X Vaporesso
    'luxe x': 'cartouche luxe x',
    # Pixo Aspire
    'pixo': 'cartouche pixo',
    # Soul GeekVape
    'soul': 'cartouche geekvape soul',
    # Ursa Nano LostVape
    'ursa nano': 'cartouche ursa nano',
    'ursa v3': 'cartouche ursa nano',
    'ursa v2': 'cartouche ursa nano',
    # UB Max LostVape
    'ub max': 'resistances ub max',
    # GTX Vaporesso
    'gtx': 'resistance gtx',
    # Veynom Air vide
    'veynom air': 'cartouche vide veynom aspire',
}

def extract_ohm(s):
    """Extrait la valeur ohm d'une chaîne (ex: 0.4, 0,4, 0.4?, 0.4 ohm)"""
    s = s.lower().replace(',','.')
    m = re.search(r'0\.(\d{1,2})\s*(?:ohm|\?|$)', s)
    if m: return f"0.{m.group(1)}"
    m = re.search(r'1\.(\d{1,2})\s*(?:ohm|\?)', s)
    if m: return f"1.{m.group(1)}"
    m = re.search(r'(\d+\.\d+)\s*(?:ohm|\?)', s)
    if m: return m.group(1)
    return None

def find_resistance_match(produit):
    """Cherche la résistance/cartouche correspondante dans Cash Mag en matchant sur l'OHM."""
    prod_low = produit.lower()
    ohm = extract_ohm(produit)
    
    for keyword, cm_prefix in RESISTANCE_MAP.items():
        if keyword in prod_low:
            # Chercher dans le catalogue avec ce préfixe + cet ohm
            candidates = []
            for p in CATALOGUE:
                lib = p['libelle'].lower()
                if cm_prefix in lib:
                    if ohm:
                        # Normaliser l'ohm du catalogue (0,15 → 0.15)
                        lib_norm = lib.replace(',','.')
                        ohm_clean = ohm.replace('.','').lstrip('0') or '0'
                        # Chercher l'ohm dans le libellé Cash Mag
                        if ohm in lib_norm or ohm.replace('.','').replace(',','') in lib_norm.replace('.','').replace(',',''):
                            candidates.append(p)
                    else:
                        candidates.append(p)
            if candidates:
                return candidates[0], 0.99
    return None, 0

print(f"Catalogue: {len(CATALOGUE)} produits | Index: {len(INDEX)} mots")

def find_best(produit, nic):
    prod_norm = normalize(produit)

    # ── Résistances et cartouches : matching par OHM ──────────────────────────
    res_match, res_score = find_resistance_match(produit)
    if res_match:
        return res_match, res_score

    # ── Correspondances manuelles ─────────────────────────────────────────────
    for cle, valeur in MANUELS.items():
        if cle in prod_norm or prod_norm in cle:
            for p in CATALOGUE:
                if valeur in normalize(p['libelle']):
                    vol_f = re.search(r'(\d+)\s*ml', produit.lower())
                    vol_c = re.search(r'(\d+)\s*ml', p['libelle'].lower())
                    if vol_f and vol_c and vol_f.group(1) == vol_c.group(1):
                        return p, 0.99
                    elif not vol_f:
                        return p, 0.95
            break

    # ── Cartouches Luxe X : matcher sur l'ohm exact ───────────────────────────
    if 'luxe x' in prod_norm and 'cartouche' in prod_norm:
        ohm_m = re.search(r'0[.,](\d)', produit)
        if ohm_m:
            ohm = ohm_m.group(1)
            for p in CATALOGUE:
                lib = normalize(p['libelle'])
                if f'0.{ohm}ohm' in lib and 'luxe x' in lib and 'cartouche' in lib:
                    return p, 0.99

    # ── Packs Pulp 60ml → chercher 50ml dans Cash Mag ────────────────────────
    if 'pulp' in prod_norm and '60ml' in produit.lower():
        produit_mod = re.sub(r'60\s*ml', '50ml', produit, flags=re.I)
        prod_norm = normalize(produit_mod)

    # ── Algorithme général ────────────────────────────────────────────────────
    qn_raw = prod_norm
    qn_raw = re.sub(r'^pack\s+|^concentre\s+', '', qn_raw)
    qn_raw = re.sub(r'\s*par\s+\d+\b', '', qn_raw)
    qn_raw = qn_raw.replace('aromes et liquides','a&l').replace('aromes liquides','a&l')
    qn_raw = re.sub(r'\s+', ' ', qn_raw).strip()
    vol_fourn = extract_volume(produit)
    nic_str = str(int(nic)) + 'mg' if nic and nic not in ('0','00') else ''
    name_words = get_name_words(produit)
    all_words = [w for w in qn_raw.split() if len(w) > 2]
    candidates = {}
    for w in all_words + name_words:
        for idx in INDEX.get(w, []):
            candidates[idx] = candidates.get(idx, 0) + 1
    if not candidates:
        candidates = {i: 0 for i in range(len(CATALOGUE))}
    top = sorted(candidates.items(), key=lambda x: -x[1])[:300]
    best, best_score = None, -1
    for idx, _ in top:
        p = CATALOGUE[idx]
        lib = normalize(p['libelle'])
        lib_al = lib.replace('aromes et liquides','a&l')
        s = max(SequenceMatcher(None, qn_raw, lib).ratio(),
                SequenceMatcher(None, qn_raw, lib_al).ratio())
        if nic_str:
            if nic_str in lib: s += 0.25
            else: s -= 0.20
        else:
            if re.search(r'\d+mg', lib) and '0mg' not in lib: s -= 0.15
        s += sum(0.04 for w in all_words if w in lib)
        name_hits = sum(1 for w in name_words if w in lib)
        s += name_hits * 0.50
        if name_words and name_hits == 0: s -= 0.60
        if vol_fourn:
            vol_cm = extract_volume(p['libelle'])
            if vol_cm:
                if vol_cm == vol_fourn: s += 0.30
                elif abs(vol_cm - vol_fourn) <= 5: s += 0.05
                else: s -= 0.30
            else: s -= 0.10
        if s > best_score:
            best_score, best = s, p
    return best, round(best_score, 2)


def parse_bl(text):
    items, seen = [], set()
    lines = text.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        m = re.match(r'(#REF\d+-\d+)(.*)', line)
        if m:
            ref = m.group(1)
            desc_parts = [m.group(2).strip()]
            j = i + 1
            while j < len(lines):
                nl = lines[j].strip()
                if re.match(r'#REF\d+-\d+', nl) or re.match(r'^Page\s*:', nl) or 'Colisage' in nl:
                    break
                if re.match(r'^\d{1,3}$', nl) and j > i + 1:
                    break
                desc_parts.append(nl)
                j += 1
            block = ' '.join(desc_parts)
            qty_m = re.search(r'\)\s*(\d{1,3})\s*$', block) or re.search(r'\)\s+(\d{1,3})', block)
            qty = int(qty_m.group(1)) if qty_m else 0
            if not qty and j < len(lines) and re.match(r'^\d{1,3}$', lines[j].strip()):
                qty = int(lines[j].strip()); j += 1
            nic_m = re.search(r'[Dd]osage\s+[Nn]icotine\s*:\s*(\d+)\s*mg', block)
            nic = str(int(nic_m.group(1))) if nic_m else '0'
            produit = re.sub(r'\s*\(.*', '', block).strip()
            produit = re.sub(r'\s*-\s*(0mg|\d+mg)\s*$', '', produit, flags=re.I).strip()
            key = ref + '|' + nic
            if qty > 0 and key not in seen:
                seen.add(key)
                items.append({'ref': ref, 'produit': produit, 'nic': nic, 'qty': qty})
            i = j
        else:
            i += 1
    return items

def parse_lca(text):
    items, seen = [], set()
    lines = [l.strip() for l in text.split('\n')]
    i = 0
    while i < len(lines):
        if re.match(r'^#REF\d+-\d+$', lines[i]):
            ref = lines[i]
            desc_parts = []
            j = i + 1
            while j < len(lines):
                nl = lines[j]
                if re.match(r'^#REF\d+-\d+$', nl): break
                if re.match(r'^\d+$', nl) and j > i + 1: break
                if any(x in nl for x in ['Sous-total','RESERVE','Aucun','IBAN','BIC','Base HT','Date d']): break
                desc_parts.append(nl)
                j += 1
            desc = ' '.join(desc_parts)
            qty = 0
            if j < len(lines) and re.match(r'^\d+$', lines[j]):
                qty = int(lines[j])
            nic_m = re.search(r'[Nn]icotine\s*:\s*(\d+)\s*mg', desc)
            nic = str(int(nic_m.group(1))) if nic_m else '0'
            produit = re.sub(r'\s*-\s*[Dd]osage\s+[Nn]icotine.*', '', desc)
            produit = re.sub(r'\s*-\s*[Cc]ontenance.*', '', produit)
            produit = re.sub(r'\s*-\s*[Cc]ouleur.*', '', produit)
            produit = re.sub(r'\s*\([^)]*\)', '', produit).strip()
            if qty > 0 and 'PLV' not in produit:
                key = ref + '|' + nic
                if key not in seen:
                    seen.add(key)
                    items.append({'ref': ref, 'produit': produit, 'nic': nic, 'qty': qty})
            i = j + 1
        else:
            i += 1
    return items

def parse_lvp(text):
    has_tva_col = 'Code\nTVA' in text or 'Code TVA' in text
    items, seen = [], set()
    lines = [l.strip() for l in text.split('\n')]
    SKIP = {'Référence','Désignation','Quantité','PU HT','Montant HT','Code',
            'TVA','Sous-total HT','INCOTERM DAP','TOTAL','Base HT','Taux TVA',
            'Montant TVA','Date','Mode','Total','Remise','IBAN','BIC','RESERVE','LVP DISTRIBUTION'}

    def is_ref_lvp(s):
        if len(s) < 4: return False
        if not re.match(r'^[A-Z0-9][A-Z0-9\-\.]{3,}$', s): return False
        if re.match(r'^\d+[\.,]\d+$', s): return False
        if s in SKIP: return False
        return True

    i = 0
    while i < len(lines):
        line = lines[i]
        if not is_ref_lvp(line):
            i += 1; continue
        ref = line
        j = i + 1
        desc_parts = []
        while j < len(lines):
            nl = lines[j]
            if is_ref_lvp(nl): break
            if re.match(r'^\d{1,2}$', nl): break
            if re.match(r'^\d+[\.,]\d+$', nl): break
            if any(w in nl.lower() for w in ['sous-total','reserve','incoterm','iban']): break
            if nl: desc_parts.append(nl)
            j += 1
        desc = ' '.join(desc_parts)
        nic_m = re.search(r'[Nn]icotine\s*[:\(]\s*(\d+)\s*mg', desc)
        if nic_m:
            nic = str(int(nic_m.group(1)))
        else:
            nic = '0'
            for k in range(i+1, min(i+8, len(lines))):
                nm = re.match(r'^(\d+)mg$', lines[k], re.I)
                if nm: nic = str(int(nm.group(1))); break
        qty = 0
        while j < len(lines):
            nl = lines[j]
            if re.match(r'^\d+$', nl):
                candidate = int(nl)
                if has_tva_col and candidate <= 2 and j+1 < len(lines) and re.match(r'^\d+$', lines[j+1]):
                    j += 1; qty = int(lines[j])
                else:
                    qty = candidate
                break
            if is_ref_lvp(nl): break
            if re.match(r'^\d+[\.,]\d+$', nl): break
            j += 1
        produit = re.sub(r'\s*\([^)]*\)', '', desc)
        produit = re.sub(r'\s*-\s*[Nn]icotine.*', '', produit)
        produit = re.sub(r'\s*-\s*[Cc]ouleur.*', '', produit)
        produit = re.sub(r'\s*-\s*[Cc]ontenance.*', '', produit)
        produit = re.sub(r'\s*-\s*[Oo]hm.*', '', produit)
        produit = re.sub(r'\s*-\s*[Vv]aleur.*', '', produit)
        produit = re.sub(r'\s*-\s*[Vv]ersion.*', '', produit)
        produit = re.sub(r'\s+(0mg|\d+mg)$', '', produit)
        produit = re.sub(r'\s+', ' ', produit).strip()
        skip_words = ['offert','base 1l','booster 10ml','fiole','bobine','film',
                      'chargeur','rouleau','echantillon','portes centaurus','offerte']
        if qty > 0 and len(produit) > 5 and not any(w in produit.lower() for w in skip_words):
            key = ref + '|' + nic
            if key not in seen:
                seen.add(key)
                items.append({'ref': ref, 'produit': produit, 'nic': nic, 'qty': qty})
        i = j + 1
    return items

def parse_greenvillage(text):
    items, seen = [], set()
    lines = [l.strip() for l in text.split('\n')]
    HEADER_STOP = ['Adresse','ROMAIN','Vapo','Centre','France','Customer',
                   'Référence','Produit','Taux','Prix','Total','Détail',
                   'Greenvillage','Pour toute','Numéro Siret','Powered']
    i = 0
    while i < len(lines):
        line = lines[i]
        m = re.match(r'^(\d[A-Z0-9]{2,}[\-]?)$', line)
        if not m:
            i += 1; continue
        ref_part1 = line
        j = i + 1
        ref_part2 = ''
        if j < len(lines) and re.match(r'^[A-Z0-9]{3,}$', lines[j]) and lines[j] != '20 %':
            ref_part2 = lines[j]; j += 1
        ref = ref_part1 + ref_part2
        # Vérifier qu'on n'est pas dans l'en-tête
        if any(h in ' '.join(lines[max(0,i-3):i]) for h in HEADER_STOP):
            i += 1; continue
        desc_parts = []
        while j < len(lines):
            nl = lines[j]
            if nl == '20 %': break
            if re.match(r'^\d[A-Z0-9]{2,}[\-]?$', nl): break
            if any(x in nl for x in ['Total produits','Détail','Greenvillage','Powered']): break
            desc_parts.append(nl); j += 1
        desc = ' '.join(desc_parts)
        if j < len(lines) and lines[j] == '20 %':
            j += 3
        qty = 0
        if j < len(lines) and re.match(r'^\d+$', lines[j]):
            qty = int(lines[j]); j += 1
        nic_m = re.search(r'[Dd]éclinaison\s*:\s*(\d+)\s*mg', desc)
        if not nic_m: nic_m = re.search(r'(\d+)\s*mg', desc)
        if nic_m:
            nic_val = int(nic_m.group(1))
            nic = '0' if nic_val == 0 else str(nic_val)
        else:
            nic = '0'
        produit = re.sub(r'\s*-\s*[Dd]éclinaison.*', '', desc)
        produit = re.sub(r'\s*\([^)]*\)', '', produit)
        produit = re.sub(r'\s*\(Par \d+\)', '', produit, flags=re.I)
        produit = produit.strip()
        skip = ['adresse','romain','vapo','centre','france','customer',
                'tank en pyrex','tube','pyrex','offert','gratuit']
        if qty > 0 and len(produit) > 5 and not any(s in produit.lower() for s in skip):
            key = ref + '|' + nic
            if key not in seen:
                seen.add(key)
                items.append({'ref': ref, 'produit': produit, 'nic': nic, 'qty': qty})
        i = j
    return items

def parse_facture_standard(text):
    items = []
    for m in re.finditer(
        r'([A-Z0-9]{6,})\s+(.+?)(?:Nicotine\s*:\s*(\d+)\s*mg[^)]*\)?\s*)?(?:0\s*%|20\s*%)\s+[\d,]+\s*\u20ac\s+(\d+)',
        text, re.I):
        ref = m.group(1)
        produit = re.sub(r'\([^)]+\)', '', m.group(2)).strip()
        produit = re.sub(r'\s*-\s*(Pulp\s*-\s*FRC|FR)\s*', ' ', produit, flags=re.I).strip()
        nic = m.group(3) or '0'
        qty = int(m.group(4))
        if qty > 0 and len(ref) >= 6:
            items.append({'ref': ref, 'produit': produit, 'nic': nic, 'qty': qty})
    return items


def parse_gfc(text):
    """GFC Provap : ref GFC..., désignation, code barre (long chiffre), qty, PU, total"""
    items, seen = [], set()
    lines = [l.strip() for l in text.split('\n')]
    i = 0
    while i < len(lines):
        line = lines[i]
        if not re.match(r'^GFC\d+', line):
            i += 1; continue
        ref = line
        j = i + 1
        desc_parts = []
        while j < len(lines):
            nl = lines[j]
            if re.match(r'^GFC\d+', nl): break
            if re.match(r'^\d{8,}$', nl): break
            if any(x in nl for x in ['Sous-total','RESERVE','Base HT','Total HT','GFC Provap']): break
            desc_parts.append(nl)
            j += 1
        desc = ' '.join(desc_parts)
        if j < len(lines) and re.match(r'^\d{8,}$', lines[j]):
            j += 1
        qty = 0
        if j < len(lines) and re.match(r'^\d+$', lines[j]):
            qty = int(lines[j]); j += 1
        nic_m = re.search(r'[Nn]icotine\s*:\s*(\d+)\s*mg', desc)
        nic = str(int(nic_m.group(1))) if nic_m else '0'
        produit = re.sub(r'\s*-\s*[Mm]od[eè]le.*', '', desc)
        produit = re.sub(r'\s*-\s*[Vv]aleur.*', '', produit)
        produit = re.sub(r'\s*-\s*[Cc]ouleur.*', '', produit)
        produit = re.sub(r'\s*\([^)]*\)', '', produit)
        produit = re.sub(r'\s*\[FID\].*', '', produit).strip()
        skip = ['fid','ornement','accu vtc','accu 50s','accu 18650','accu 21700']
        if qty > 0 and len(produit) > 3 and not any(s in produit.lower() for s in skip):
            key = ref + '|' + nic
            if key not in seen:
                seen.add(key)
                items.append({'ref': ref, 'produit': produit, 'nic': nic, 'qty': qty})
        i = j
    return items

def parse_adns(text):
    """ADNS : description, Origine UE/HORS UE, qty, prix, total"""
    items, seen = [], set()
    lines = [l.strip() for l in text.split('\n')]
    i = 0
    while i < len(lines):
        line = lines[i]
        if i+1 < len(lines) and lines[i+1] in ('UE', 'HORS UE') and len(line) > 5:
            desc = line
            j = i + 2
            qty = 0
            if j < len(lines) and re.match(r'^\d+$', lines[j]):
                qty = int(lines[j]); j += 1
            nic_m = re.search(r'[Dd]osage\s+nicotine\s*:\s*0*(\d+)\s*mg', desc)
            nic = str(int(nic_m.group(1))) if nic_m else '0'
            produit = re.sub(r'\s*-\s*[Dd]osage\s+nicotine.*', '', desc)
            produit = re.sub(r'\s*-\s*[Cc]ouleur.*', '', produit)
            produit = re.sub(r'\s*-\s*[Ii]ntensit.*', '', produit)
            produit = re.sub(r'\s*\([^)]*\)', '', produit).strip()
            skip = ['echantillon','flyer','reduction','goodies']
            if qty > 0 and len(produit) > 3 and not any(s in produit.lower() for s in skip):
                key = desc[:30] + '|' + nic
                if key not in seen:
                    seen.add(key)
                    items.append({'ref': 'ADNS', 'produit': produit, 'nic': nic, 'qty': qty})
            i = j
        else:
            i += 1
    return items


def parse_grossiste(text):
    """Grossiste Ecigarette : ref ARxxxxx, designation multi-lignes, 20%, prix, qty, total"""
    items, seen = [], set()
    lines = [l.strip() for l in text.split('\n')]
    i = 0
    while i < len(lines):
        line = lines[i]
        if not re.match(r'^AR\d{5,}$', line):
            i += 1; continue
        ref = line
        j = i + 1
        desc_parts = []
        while j < len(lines):
            nl = lines[j]
            if nl == '20 %': break
            if re.match(r'^AR\d{5,}$', nl): break
            if any(x in nl for x in ['Reductions','Détail','Grossiste','Powered']): break
            desc_parts.append(nl)
            j += 1
        desc = ' '.join(desc_parts)
        if j < len(lines) and lines[j] == '20 %':
            j += 1
            if j < len(lines) and (lines[j] == '--' or '\u20ac' in lines[j]): j += 1
            if j < len(lines) and '\u20ac' in lines[j]: j += 1
        qty = 0
        if j < len(lines) and re.match(r'^\d+$', lines[j]):
            qty = int(lines[j]); j += 1
        nic_m = re.search(r'[Dd]osage\s*:\s*(\d+)\s*mg', desc)
        nic = str(int(nic_m.group(1))) if nic_m else '0'
        produit = re.sub(r'\s*-\s*[Dd]osage\s*:.*', '', desc)
        produit = re.sub(r'\s*-\s*[Ss]aveur\s*:.*', '', produit)
        produit = re.sub(r'\s*-\s*[Cc]ouleur\s*:.*', '', produit)
        produit = re.sub(r'\s*\([Bb]oite\s+de\s+\d+\)', '', produit)
        produit = re.sub(r'\s*\([^)]*\)', '', produit)
        produit = re.sub(r'^E liquide\s+', '', produit, flags=re.I)
        produit = produit.strip()
        saveur_m = re.search(r'[Ss]aveur\s*:\s*(.+?)(?:\s*$|\s*-)', desc)
        if saveur_m:
            produit = produit + ' ' + saveur_m.group(1).strip()
        skip = ['reductions','offerte','livraison','1 boite']
        if qty > 0 and len(produit) > 3 and not any(s in produit.lower() for s in skip):
            key = ref + '|' + nic
            if key not in seen:
                seen.add(key)
                items.append({'ref': ref, 'produit': produit, 'nic': nic, 'qty': qty})
        i = j
    return items

def parse_doc(text):
    # Détecter le fournisseur
    if 'greenvillage' in text.lower():
        return parse_greenvillage(text)
    if 'Grossiste Ecigarette' in text or 'grossiste-ecigarette' in text.lower():
        return parse_grossiste(text)
    if 'GFC Provap' in text or 'gfc-provap' in text.lower():
        return parse_gfc(text)
    if 'ADNS' in text and 'Vente en gros' in text:
        return parse_adns(text)
    if 'LCA DISTRIBUTION' in text:
        return parse_lca(text)
    if 'LVP DISTRIBUTION' in text:
        return parse_lvp(text)
    if '#REF' in text and ('Colisage' in text or 'Dosage Nicotine' in text):
        return parse_bl(text)
    return parse_facture_standard(text)

# ── Génération Excel ───────────────────────────────────────────────────────────

def generate_excel(results, filename):
    VERT = "C6EFCE"; VERT_TXT = "276221"
    ORANGE = "FFEB9C"; ORANGE_TXT = "9C5700"
    BLEU = "1F4E79"; BLEU2 = "2E75B6"
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entree stock"
    ws.merge_cells('A1:H1')
    ws['A1'] = f"ENTREE DE STOCK - {filename}"
    ws['A1'].font = Font(bold=True, size=12, color='FFFFFF', name='Arial')
    ws['A1'].fill = PatternFill('solid', start_color=BLEU)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22
    ok_c = sum(1 for r in results if r['statut'] == 'OK')
    ws.merge_cells('A2:H2')
    ws['A2'] = f"{len(results)} references | OK: {ok_c} | A verifier: {len(results)-ok_c}"
    ws['A2'].font = Font(size=10, color='595959', name='Arial')
    ws['A2'].fill = PatternFill('solid', start_color='DEEAF1')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 15
    for col, h in enumerate(['#','Ref Fournisseur','Produit Fournisseur','Nicotine','Qte',
                              'Libelle Cash Mag','ID Cash Mag','Statut'], 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill = PatternFill('solid', start_color=BLEU2)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.row_dimensions[3].height = 18
    for i, r in enumerate(results, 1):
        rn = i + 3
        is_av = r['statut'] == 'A VERIFIER'
        nic_label = (r['nic'] + 'mg') if r['nic'] not in ('0','00') else '0mg'
        for col, val in enumerate([i, r['ref'], r['produit'], nic_label, r['qty'],
                                    r['cashMagLibelle'], r['cashMagId'], r['statut']], 1):
            c = ws.cell(row=rn, column=col, value=val)
            c.font = Font(name='Arial', size=10)
            c.border = border
            c.alignment = Alignment(vertical='center',
                                    horizontal='center' if col in [1,4,5,7,8] else 'left')
            if col == 8:
                c.fill = PatternFill('solid', start_color=ORANGE if is_av else VERT)
                c.font = Font(name='Arial', size=10, bold=True,
                              color=ORANGE_TXT if is_av else VERT_TXT)
        ws.row_dimensions[rn].height = 16
    for col, w in zip('ABCDEFGH', [4, 18, 34, 10, 6, 40, 13, 13]):
        ws.column_dimensions[col].width = w
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ── Routes ─────────────────────────────────────────────────────────────────────

# Stockage temporaire des résultats en mémoire
import uuid
RESULTATS_CACHE = {}

@app.route('/')
def index():
    return render_template('index.html', nb_produits=len(CATALOGUE))

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier recu'}), 400
    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Envoyez un fichier PDF'}), 400
    try:
        doc = fitz.open(stream=file.read(), filetype='pdf')
        text = "\n".join(page.get_text() for page in doc)
        items = parse_doc(text)
        if not items:
            return jsonify({'error': 'Aucune reference trouvee dans ce PDF. Fournisseur non supporte ?'}), 400
        results = []
        for item in items:
            match, score = find_best(item['produit'], item['nic'])
            results.append({**item,
                'cashMagLibelle': match['libelle'] if match else 'NON TROUVE',
                'cashMagId': str(match['id']) if match else '',
                'score': score,
                'statut': 'OK' if score >= 0.65 else 'A VERIFIER'})
        fname = os.path.splitext(file.filename)[0]
        # Stocker en cache pour téléchargement ultérieur
        cache_id = str(uuid.uuid4())
        RESULTATS_CACHE[cache_id] = {'results': results, 'fname': fname}
        ok = sum(1 for r in results if r['statut'] == 'OK')
        av = len(results) - ok
        return jsonify({
            'cache_id': cache_id,
            'fname': fname,
            'total': len(results),
            'ok': ok,
            'av': av,
            'results': results
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download/<cache_id>')
def download(cache_id):
    if cache_id not in RESULTATS_CACHE:
        return jsonify({'error': 'Session expirée, retraitez le PDF'}), 404
    data = RESULTATS_CACHE[cache_id]
    excel = generate_excel(data['results'], data['fname'])
    return send_file(excel, as_attachment=True,
        download_name=f'entree_stock_{data["fname"]}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
